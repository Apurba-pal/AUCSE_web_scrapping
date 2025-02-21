import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import pandas as pd
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects

URL = "https://www.aucse.in/people/student/btech/cse-batch-2022-2026"
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36"
}

# Send GET request with error handling
try:
    response = requests.get(URL, headers=headers, timeout=10)
    response.raise_for_status()
except (ConnectionError, Timeout, TooManyRedirects) as e:
    print(f"Error fetching the URL: {e}")
    exit(1)

soup = BeautifulSoup(response.text, "html.parser")

# Create an Excel workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Scraped Data"
sheet.append(["name", "student_id", "roll_no", "sec", "aoi", "phone", "email", "course", "year", "status"])

# Extracting data from HTML
sections = soup.find_all("section", {"class": ["yaqOZd", "cJgDec", "tpmmCb"]})
for section in sections:
    paragraphs = section.find_all("p", class_="zfr3Q CDt4Ke")
    sl_no = paragraphs[0].text.strip() if len(paragraphs) > 0 else "N/A"
    roll_number = paragraphs[1].text.strip() if len(paragraphs) > 1 else "N/A"
    name = paragraphs[2].text.strip() if len(paragraphs) > 2 else "N/A"
    specialization = paragraphs[3].text.strip() if len(paragraphs) > 3 else "N/A"
    area_of_interest = paragraphs[4].text.strip() if len(paragraphs) > 4 else "N/A"
    
    # Replace "NA" with "B.Tech. CSE (Core)" in Specialization
    if specialization == "NA":
        specialization = "B.Tech. CSE (Core)"
    
    batch = "2022"  
    phone_no = "00"
    status = 1
    
    sheet.append([name, "", roll_number, "", area_of_interest, phone_no, "", specialization, batch, status])

# Save the data to an Excel file
wb.save("college_data.xlsx")

# Load the CSV files
csv_files = [
    "B Tech CSE6th Sem Sec-B_final.xlsx",
    "SecA_Student's information 22-23.xlsx",
    "Student list_SEC C.xlsx"
]

# Create a mapping of Roll Number to Registration No, Section, Phone Number, and Email ID
roll_to_reg_section_phone_email = {}
for csv_file in csv_files:
    df = pd.read_excel(csv_file)
    print(f"Columns in {csv_file}: {df.columns.tolist()}")  # Print columns for debugging
    for _, row in df.iterrows():
        email_id = row['Email ID'] if 'Email ID' in df.columns else row['Email ID '] if 'Email ID ' in df.columns else ""
        roll_to_reg_section_phone_email[row['Roll Number']] = (row['Registration No'], row['Section'], row['Phone Number'], email_id)

# Load the college_data.xlsx file
wb = load_workbook("college_data.xlsx")
sheet = wb.active

# Update the Registration No, Section, Phone Number, and Email ID for each Roll Number
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    roll_number = row[2].value
    if roll_number in roll_to_reg_section_phone_email:
        row[1].value = roll_to_reg_section_phone_email[roll_number][0]
        row[3].value = roll_to_reg_section_phone_email[roll_number][1]
        row[5].value = roll_to_reg_section_phone_email[roll_number][2]
        row[6].value = roll_to_reg_section_phone_email[roll_number][3]

# Save the updated Excel file
wb.save("college_data_updated.xlsx")
print("Data successfully scraped and saved to college_data_updated.xlsx")
